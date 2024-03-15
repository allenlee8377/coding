#把已經合併完成的excel，各項運算，找出周漲跌機率
#可以輸入特定的範圍，在vlook的時候就處理好


from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# 定义运算函数
def calculate(value):
    try:
        value = float(value)  # 尝试将值转换为浮点数
        if value > 0:
            return 1
        elif value < 0:
            return -1
        else:
            return 0
    except ValueError:
        return "N/A"  # 如果值无法转换为浮点数，返回 "N/A"

# 读取原始 Excel 文件
original_file_path = r'C:\Users\allenlee\Desktop\DBtest\excel_all\new_9962_A.xlsx'
wb_original = load_workbook(original_file_path)

# 创建新的 Excel 文件
wb_new = Workbook()

# 获取当前年份的最后两位数
current_year = str(datetime.now().year % 100)

# 添加代码以接收用户输入的年份范围
year_range = input("请輸入年份（例如：23,22,20,15）：").split(",")
year_range = [int(year.strip()) for year in year_range]

# 处理每个分页
for sheet_name in wb_original.sheetnames:
    ws_original = wb_original[sheet_name]
    ws_new = wb_new.create_sheet(title=sheet_name)

    # 在第8列中填入 "漲跌統計"
    ws_original['H1'] = "漲跌統計"

    # 在H列作运算，并存储到新的第8列
    for row in range(1, ws_original.max_row + 1):
        g_value = ws_original.cell(row=row, column=7).value
        h_value = calculate(g_value)
        ws_original.cell(row=row, column=8, value=h_value)

    # 将第1列和第8列的数据存储到新文件的第1列和第2列，并清空第8列的数据
    for row in range(1, ws_original.max_row + 1):
        value_1 = ws_original.cell(row=row, column=1).value
        value_8 = ws_original.cell(row=row, column=8).value
        ws_new.cell(row=row, column=1, value=value_1)
        ws_new.cell(row=row, column=2, value=value_8)
        #ws_new.cell(row=row, column=8, value=None)

    # 在新文件中添加 "數列" 和 "W53" 到 "W01"
    ws_new['H1'] = sheet_name
    ws_new['H2'] = "週期"
    for row, week_number in enumerate(range(53, 0, -1), start=3):
        ws_new.cell(row=row, column=8, value=f'W{week_number:02}')

    # 在新文件中添加当前年份的最后两位数到-11
    for i, year in enumerate(range(int(current_year), int(current_year) - 11, -1), start=9):
        ws_new.cell(row=2, column=i, value=str(year))

    # 根据用户输入的年份范围进行计算
    for year in year_range:
        # 根据VLOOKUP的方式填充数据
        for col_offset, year_value in enumerate(range(int(current_year), int(current_year) - 11, -1), start=9):
            if year_value == year:
                for row in range(3, 56):
                    week_number = ws_new.cell(row=row, column=8).value
                    if week_number:
                        lookup_value = f"{year_value}{week_number}"
                        value_to_fill = None
                        for i, cell in enumerate(ws_original['A'], start=1):
                            if cell.value == lookup_value:
                                value_to_fill = ws_new.cell(row=i, column=2).value
                                break
                        ws_new.cell(row=row, column=col_offset, value=value_to_fill if value_to_fill else "N/A")

    # 在 C2 到 F2 中填写固定的数值
    fixed_values = [1, 0, -1, "N/A"]
    for col_offset, value in enumerate(fixed_values, start=3):
        ws_new.cell(row=2, column=col_offset, value=value)

    # 计算每一行中特定数值的数量
    for row in range(3, 56):
        for col_offset, value in enumerate([1, 0, -1, "N/A"], start=3):
            count = sum(1 for col in range(9, 19) if ws_new.cell(row=row, column=col).value == value)
            ws_new.cell(row=row, column=col_offset, value=count)

    # 在 G2 中填写 "漲跌機率"
    ws_new['G1'] = sheet_name
    ws_new['G2'] = "漲跌機率"

    # 计算漲跌機率
    for row in range(3, 56):
        total_count = sum(ws_new.cell(row=row, column=col_offset).value for col_offset in range(3, 7))
        probability = round(ws_new.cell(row=row, column=3).value / total_count * 100, 2) if total_count != 0 else 0
        ws_new.cell(row=row, column=7, value=probability)

# 删除默认创建的 Sheet
wb_new.remove(wb_new["Sheet"])

# 保存新文件
new_file_path = original_file_path.replace('.xlsx', '_modified.xlsx')
wb_new.save(new_file_path)
