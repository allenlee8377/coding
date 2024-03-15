import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 指定包含 HTML 檔案的目錄路徑
html_files_directory = 'C:\\Users\\allenlee\\Desktop\\DBtest\\html_A'

# 選擇要添加底色的欄位的列和欄索引
highlight_column_index = 7  # 假設要添加底色的欄位是第二列
highlight_color = 'FFFF00'  # 黃色

# 設置底色樣式
fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

# 讀取並處理每個 HTML 檔案
for filename in os.listdir(html_files_directory):
    if filename.endswith('.html'):
        # 提取檔案名稱作為 Excel 檔案名稱
        excel_filename = filename.replace('.html', '.xlsx')

        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active

        # 讀取 HTML 檔案內容
        with open(os.path.join(html_files_directory, filename), 'r', encoding='utf-8') as file:
            html_content = file.read()

        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        # 獲取所有表格標籤
        tables = soup.find_all('table')
        
        # 將每個表格的內容寫入工作表的不同區域
        for idx, table in enumerate(tables, start=1):
            # 計算每個表格的起始列和欄
            start_row = 1
            start_col = 1
            
            # 將表格的每個儲存格寫入工作表
            for row_idx, row in enumerate(table.find_all('tr'), start=start_row):
                for col_idx, cell in enumerate(row.find_all(['td', 'th']), start=start_col):
                    # 寫入單元格內容
                    ws.cell(row=row_idx, column=col_idx, value=cell.get_text())
                    # 如果是指定的欄位，則設置底色
                    if col_idx == highlight_column_index:
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        # 儲存 Excel 檔案到指定路徑
        excel_file_path = os.path.join(html_files_directory, excel_filename)
        wb.save(excel_file_path)
        print("HTML 檔案", filename, "已成功轉換為 Excel 檔案:", excel_filename, "並儲存到路徑:", html_files_directory)
