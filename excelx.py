#如何使用套件
# $ pip install openpyxl  安裝這個套件
# 物件: 可以大寫，function不會有大寫
# py內都是物件，只是型別(type)不一樣
# 5.3 = int / flat
# Workbook >>  他的type就是 Workbook



from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
ws['B1'] = 'Allen'

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([4, 5, 6])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

