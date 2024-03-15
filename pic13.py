#可以輸入檔名

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from os.path import join, isfile

# 定義運算函數
def calculate(value):
    try:
        value = float(value)  # 尝试将值转换为浮点数
        if value > 0:
            return 1
        elif value < 0:
            return -1
        else:
            return 0
    except ValueError:
        return "N/A"  # 如果值无法转换为浮点数，返回 "N/A"

# 預設的資料夾路徑
default_folder_path = r'C:\K'

# 提示用戶輸入檔案名稱
file_name = input("請輸入要處理的Excel檔案名稱（不需要附檔名）：")

# 獲取完整的檔案路徑，假設附檔名為 ".xlsx"
file_path = join(default_folder_path, file_name + ".xlsx")

# 檢查檔案是否存在，如果存在就讀取它
if isfile(file_path):
    wb_original = load_workbook(file_path)
else:
    print("無此檔案。")

    # 繼續詢問是否要輸入其他檔案名稱
    while True:
        choice = input("是否要輸入其他檔案名稱？(y/n): ").strip().lower()
        if choice == 'n':
            break
        elif choice == 'y':
            file_name = input("請輸入要處理的Excel檔案名稱（不需要附檔名）：")
            file_path = join(default_folder_path, file_name + ".xlsx")
            if isfile(file_path):
                wb_original = load_workbook(file_path)
                break
            else:
                print("無此檔案。")
        else:
            print("請輸入有效的選項。")

# 創建新的 Excel 文件
wb_new = Workbook()

# 獲取當前年份的最后兩位数
current_year = str(datetime.now().year % 100)

# 添加代碼以接收用戶輸入的年份範圍
year_range = input("請輸入年份（例如：23,22,20,15）：").split(",")
year_range = [int(year.strip()) for year in year_range]

# 處理每個分頁
for sheet_name in wb_original.sheetnames:
    ws_original = wb_original[sheet_name]
    ws_new = wb_new.create_sheet(title=sheet_name)

    # 在第8列中填入 "漲跌統計"
    ws_original['H1'] = "漲跌統計"

    # 在H列作運算，并存儲到新的第8列
    for row in range(1, ws_original.max_row + 1):
        g_value = ws_original.cell(row=row, column=7).value
        h_value = calculate(g_value)
        ws_original.cell(row=row, column=8, value=h_value)

    # 將第1列和第8列的數據存儲到新文件的第1列和第2列，并清空第8列的數據
    for row in range(1, ws_original.max_row + 1):
        value_1 = ws_original.cell(row=row, column=1).value
        value_8 = ws_original.cell(row=row, column=8).value
        ws_new.cell(row=row, column=1, value=value_1)
        ws_new.cell(row=row, column=2, value=value_8)
        #ws_new.cell(row=row, column=8, value=None)

    # 在新文件中添加 "數列" 和 "W53" 到 "W01"
    ws_new['H1'] = sheet_name
    ws_new['H2'] = "週期"
    for row, week_number in enumerate(range(53, 0, -1), start=3):
        ws_new.cell(row=row, column=8, value=f'W{week_number:02}')

    # 在新文件中添加當前年份的最后兩位数到-10
    for i, year in enumerate(range(int(current_year), int(current_year) - 10, -1), start=9):
        ws_new.cell(row=2, column=i, value=str(year))

    # 根據用戶輸入的年份範圍進行計算
    for year in year_range:
        # 根據VLOOKUP的方式填充數據
        for col_offset, year_value in enumerate(range(int(current_year), int(current_year) - 10, -1), start=9):
            if year_value == year:
                for row in range(3, 56):
                    week_number = ws_new.cell(row=row, column=8).value
                    if week_number:
                        lookup_value = f"{year_value}{week_number}"
                        value_to_fill = None
                        for i, cell in enumerate(ws_original['A'], start=1):
                            if cell.value == lookup_value:
                                value_to_fill = ws_new.cell(row=i, column=2).value
                                break
                        ws_new.cell(row=row, column=col_offset, value=value_to_fill if value_to_fill else "N/A")

    # 在 C2 到 F2 中填寫固定的數值
    fixed_values = [1, 0, -1, "N/A"]
    for col_offset, value in enumerate(fixed_values, start=3):
        ws_new.cell(row=2, column=col_offset, value=value)

    # 計算每一行中特定數值的數量
    for row in range(3, 56):
        for col_offset, value in enumerate([1, 0, -1, "N/A"], start=3):
            count = sum(1 for col in range(9, 19) if ws_new.cell(row=row, column=col).value == value)
            ws_new.cell(row=row, column=col_offset, value=count)

    # 在 G2 中填寫 "漲跌機率"
    ws_new['G1'] = sheet_name
    ws_new['G2'] = "漲跌機率"

    # 計算漲跌機率
    for row in range(3, 56):
        total_count = sum(ws_new.cell(row=row, column=col_offset).value for col_offset in range(3, 7))
        probability = round(ws_new.cell(row=row, column=3).value / total_count * 100, 2) if total_count != 0 else 0
        ws_new.cell(row=row, column=7, value=probability)

# 刪除默認創建的 Sheet
wb_new.remove(wb_new["Sheet"])

# 保存新文件
new_file_path = file_path.replace('.xlsx', '_mod.xlsx')
wb_new.save(new_file_path)
