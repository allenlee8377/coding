#這個是統一全部都把每年份漲跌機率分析

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from os import listdir
from os.path import join, isfile
from tqdm import tqdm

# 定義運算函數
def calculate(value):
    if value is None:
        return "N/A"
    try:
        value = float(value)  # 尝试将值转换为浮点数
        if value > 0:
            return 1
        elif value < 0:
            return -1
        else:
            return 0
    except ValueError:
        return "N/A"  # 如果值无法转换为浮点数，返回 "N/A"

# 處理目標路徑
target_folder = r'C:\Users\allenlee\Desktop\DBtest\excel_all\re_do'

# 獲取目標路徑下所有檔案
files = [f for f in listdir(target_folder) if isfile(join(target_folder, f))]

# 添加代碼以接收用戶輸入的年份範圍
year_input = input("請輸入年份（例如：23,22,20,15）：")
year_range = [int(year.strip()) for year in year_input.split(",")]

# 遍歷每個 Excel 檔案
for file_name in tqdm(files, desc="Processing files", unit="file"):
    if file_name.endswith('.xlsx'):
        file_path = join(target_folder, file_name)

        # 讀取 Excel 檔案
        wb_original = load_workbook(file_path)
        
        # 創建新的 Excel 文件
        wb_new = Workbook()

        # 獲取當前年份的最后兩位数
        current_year = str(datetime.now().year % 100)

        # 處理每個分頁
        for sheet_name in wb_original.sheetnames:
            ws_original = wb_original[sheet_name]

            # 創建新的分頁，名稱為原始檔案名稱的四位數字部分
            new_sheet_name = ''.join(filter(str.isdigit, file_name))
            ws_new = wb_new.create_sheet(title=new_sheet_name)

            # 在第8列中填入 "漲跌統計"
            ws_original['H1'] = "漲跌統計"

            # 在H列作運算，并存儲到新的第8列
            for row in range(1, ws_original.max_row + 1):
                g_value = ws_original.cell(row=row, column=7).value
                h_value = calculate(g_value)
                ws_original.cell(row=row, column=8, value=h_value)

            # 將第1列和第8列的數據存儲到新文件的第1列和第2列，并清空第8列的數據
            for row in range(1, ws_original.max_row + 1):
                value_1 = ws_original.cell(row=row, column=1).value
                value_8 = ws_original.cell(row=row, column=8).value
                ws_new.cell(row=row, column=1, value=value_1)
                ws_new.cell(row=row, column=2, value=value_8)

            # 在新文件中添加 "數列" 和 "W53" 到 "W01"
            ws_new['H1'] = new_sheet_name
            ws_new['H2'] = "週期"
            for row, week_number in enumerate(range(53, 0, -1), start=3):
                ws_new.cell(row=row, column=8, value=f'W{week_number:02}')

            # 在新文件中添加当前年份的最后两位数到-10
            for i, year in enumerate(range(int(current_year), int(current_year) - 10, -1), start=9):
                ws_new.cell(row=2, column=i, value=str(year))

            # 根据用户输入的年份范围进行计算
            for year in year_range:
                # 根据VLOOKUP的方式填充数据
                for col_offset, year_value in enumerate(range(int(current_year), int(current_year) - 10, -1), start=9):
                    if year_value == year:
                        for row in range(3, 56):
                            week_number = ws_new.cell(row=row, column=8).value
                            if week_number:
                                lookup_value = f"{year_value}{week_number}"
                                value_to_fill = None
                                for i, cell in enumerate(ws_original['A'], start=1):
                                    if cell.value == lookup_value:
                                        value_to_fill = ws_new.cell(row=i, column=2).value
                                        break
                                ws_new.cell(row=row, column=col_offset, value=value_to_fill if value_to_fill else "N/A")

            # 在 C2 到 F2 中填寫固定的數值
            fixed_values = [1, 0, -1, "N/A"]
            for col_offset, value in enumerate(fixed_values, start=3):
                ws_new.cell(row=2, column=col_offset, value=value)

            # 計算每一行中特定數值的數量
            for row in range(3, 56):
                for col_offset, value in enumerate([1, 0, -1, "N/A"], start=3):
                    count = sum(1 for col in range(9, 19) if ws_new.cell(row=row, column=col).value == value)
                    ws_new.cell(row=row, column=col_offset, value=count)

            # 在 G2 中填寫 "漲跌機率"
            ws_new['G1'] = new_sheet_name
            ws_new['G2'] = "漲跌機率"

            # 計算漲跌機率
            for row in range(3, 56):
                total_count = sum(ws_new.cell(row=row, column=col_offset).value for col_offset in range(3, 7))
                probability = round(ws_new.cell(row=row, column=3).value / total_count * 100, 2) if total_count != 0 else 0
                ws_new.cell(row=row, column=7, value=probability)

        # 刪除沒有資料的分頁
        for sheet_name in wb_new.sheetnames:
            ws = wb_new[sheet_name]
            if len(ws.dimensions) == 1 and len(list(ws.values)) == 1:
                wb_new.remove(ws)

        # 保存新文件
        new_file_name = file_name.replace('_A.xlsx', '_B.xlsx')
        new_file_path = join(target_folder, new_file_name)
        wb_new.save(new_file_path)

print("所有 Excel 檔案處理完成。")
